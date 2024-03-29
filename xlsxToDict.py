#!/usr/bin/env python
# coding: utf-8


import pdb
import argparse
import pickle
import openpyxl
from pprint import pprint as ppr
from pprint import pformat as pprs
from unidecode import unidecode

def normalizeKey(String):
	toreplace={' ':'_','(':'_',')':'_'}
	resultatTmp=unidecode(String)
	strList=[]
	for c in resultatTmp:
		if c in toreplace:
			strList.append(toreplace[c]) 
		elif c.isalnum():
			strList.append(c)
	resultat=''.join( c for c in strList)
	return resultat

def testNotEmptyLine(row):
	for val in row:
		if val:
			return True
			
	return False
		
class xlsContainer(object):
	"TDB Container"
	
	def __init__(self,excel_file):

		self.headers={}
		try:
			xl_workbook = openpyxl.load_workbook(excel_file)

		except AssertionError as E:
			pdb.set_trace()
			print(E)
			
		sheet_names = xl_workbook.sheetnames
		xl_sheets={}
		self.datas={}
		self.datasRaw={}
		self.headersId={}
		self.headers={}
		
		for sheet_name in sheet_names:
			rows_val=[]
			xl_sheets[sheet_name] = xl_workbook[sheet_name]
			xl_sheet_cur=xl_sheets[sheet_name]
			rows_cur=[ row_cur for row_cur in xl_sheet_cur.iter_rows(1, xl_sheet_cur.max_row) ]
			for row in rows_cur:
				row_val=list(map(lambda y: y.value, row))
				rows_val.append(row_val)

			self.datasRaw[sheet_name]=rows_val[1:]
			self.datasRaw[sheet_name]
			self.headers[sheet_name]=rows_val[0]
			self.headersId[sheet_name]={ key:value for key,value in enumerate(self.headers[sheet_name]) }

			
		for sheet_name in sheet_names:
			self.datas[ sheet_name]=[]
			for row_val in self.datasRaw[ sheet_name]:
				dict_cur={self.headersId[sheet_name][key]:value for key,value in enumerate(row_val)}
				if not testNotEmptyLine(row_val):
					continue
				self.datas[ sheet_name].append(dict_cur)
		

				
	def save(self,filename):
		with open(filename,'wb') as file__:
			pickle.dump(self,file__)
			
	def load(self,filename):
		with open(filename,'rb') as file__:
			self=pickle.load(file__)
			
	def __str__(self):
		return pprs(self.datas)
		
	def __repr__(self):
		return pprs(self.datas)
			
if __name__ == "__main__":

	parser = argparse.ArgumentParser()
	parser.add_argument("--excel",  action="store",help="fichier excel contenant les informations",required=True)
	parser.add_argument("-s", "--save",  action="store",help="fichier de sauvegarde dump",required=False)
	parser.add_argument("-d", "--dump",  action="store",help="load dump",required=False)
	args = parser.parse_args()
	
	objFromXls=xlsContainer(args.excel)
	
	print(objFromXls)
	

		