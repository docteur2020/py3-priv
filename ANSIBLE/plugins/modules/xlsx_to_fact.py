#!/usr/bin/python

#

#	 Copyright (c) 2019 World Wide Technology, Inc. All rights reserved.

#	 GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

#

ANSIBLE_METADATA = {

	'metadata_version': '0.1',

	'status': ['preview'],

	'supported_by': 'docteur2020@yahoo.fr'

}



DOCUMENTATION = '''

---

module: xlsx_to_fact

short_description: Reads a xlsx file and returns as ansible facts a dictionaries for each sheet.
		version_added: "0.1"

description:

	- Reads the xlsx file specified and output Ansible facts in the form of a list with each

	- element in the list as a dictionary. The column header is used as the key and the contents

	- of the cell as the value. Also returns a dictionary with each column header as a key, the

	- value is a list of unique values for the column, a set.

options:

	src:

		description:

			- The CSV formatted input file

		required: true

		type: str


author:

	- myself

'''


EXAMPLES = '''

TBD

'''

import csv
from ansible.module_utils.basic import AnsibleModule


import openpyxl


import warnings
warnings.filterwarnings("ignore")
ERROR = 1

OK = 0

def normalizeKey(String):
	toreplace={' ':'_','(':'_',')':'_'}
	resultatTmp=unidecode(String)
	strList=[]
	for c in resultatTmp:
		if c in toreplace:
			strList.append(toreplace[c]) 
		elif c.isalnum():
			strList.append(c)
	resultat=''.join( c for c in strList)
	return resultat

def testNotEmptyLine(row):
	for val in row:
		if val:
			return True
			
	return False
		
class xlsContainer(object):
	"TDB Container"
	
	def __init__(self,excel_file):

		self.headers={}
		try:
			xl_workbook = openpyxl.load_workbook(excel_file)

		except AssertionError as E:
			print(E)
			
		sheet_names = xl_workbook.sheetnames
		xl_sheets={}
		self.datas={}
		self.datasRaw={}
		self.headersId={}
		self.headers={}

		for sheet_name in sheet_names:
			rows_val=[]
			xl_sheets[sheet_name] = xl_workbook[sheet_name]
			xl_sheet_cur=xl_sheets[sheet_name]
			rows_cur=[ row_cur for row_cur in xl_sheet_cur.iter_rows(1, xl_sheet_cur.max_row) ]
			for row in rows_cur:
				row_val=list(map(lambda y: y.value, row))
				rows_val.append(row_val)
			
			self.datasRaw[sheet_name]=rows_val[1:]
			self.datasRaw[sheet_name]
			self.headers[sheet_name]=rows_val[0]
			self.headersId[sheet_name]={ key:str(value).strip() for key,value in enumerate(self.headers[sheet_name]) }
		
		for sheet_name in sheet_names:
			self.datas[ sheet_name]=[]
			for row_val in self.datasRaw[ sheet_name]:
				dict_cur={self.headersId[sheet_name][key]:str(value).strip() for key,value in enumerate(row_val)}
				if not testNotEmptyLine(row_val):
					continue
				self.datas[ sheet_name].append(dict_cur)
		

def read_xlsx_dict(input_file):
	
	result = dict(
		changed=False,
		original_message='',
		message=''
	)


	
	try:

		xlsContainerObj=xlsContainer(input_file)

	except Exception as err:
		return(ERROR,str(err))
		
	result['changed'] = False
	result['msg'] = "Data sucessfully imported"
	resultat=xlsContainerObj.datas
	result["results"]=resultat
	
	return (OK,result)
		
def main():

	"""

	MAIN: manage areguments and error checking, exit or fail as appropriate

	"""

	module = AnsibleModule(argument_spec=dict(

				 src=dict(required=True, type='str'),

				 ),

				 add_file_common_args=True)


	

	code, response = read_xlsx_dict(module.params["src"])

	#if module.check_mode:
	#	module.exit_json(**result)
	#	return OK
	#	
	code=OK

	if code == ERROR:

		module.fail_json(msg=response)

	else:

		module.exit_json(meta=response)



	return code





if __name__ == '__main__':
	main()